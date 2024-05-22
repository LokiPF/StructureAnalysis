import configparser
import csv
import re

import numpy as np
import openpyxl
import pandas as pd

from Logger import Logger
from __structs__ import LoadCase, LoadsInPlane, LoadsStringers, IO, Parameters, Stringer

logger = Logger('read_excel').logger


def read_config() -> [LoadCase]:
    logger.info('Reading config file')
    config = configparser.ConfigParser()
    config.read('config.ini')
    in_plane_stresses = config['INPUT']['in_plane_stresses']
    axial_stringer_stresses = config['INPUT']['axial_stringer_stresses']
    sigma_ul = float(config['PARAMETERS']['sigma_ul'])
    io = IO(config['OUTPUT']['output_file'])
    sigma_ul = float(config['PARAMETERS']['sigma_ul'])
    sigma_yield = float(config['PARAMETERS']['sigma_yield'])
    mu = float(config['PARAMETERS']['mu'])
    sf = float(config['PARAMETERS']['sf'])
    a = float(config['PARAMETERS']['a'])
    b = float(config['PARAMETERS']['b'])
    t = float(config['PARAMETERS']['t'])
    stringer_base_w = float(config['PARAMETERS']['stringer_base_w'])
    stringer_base_t = float(config['PARAMETERS']['stringer_base_t'])
    stringer_height = float(config['PARAMETERS']['stringer_height'])
    stringer_neck_width = float(config['PARAMETERS']['stringer_neck_width'])
    main = pd.read_excel(io.output_file, sheet_name='in')
    E = float(main.iat[6, 1])
    parameters = Parameters(sigma_ul=sigma_ul, sigma_yield=sigma_yield, mu=mu, a=a, b=b, t=t,
                            stringer_base_w=stringer_base_w, stringer_height=stringer_height,
                            stringer_base_t=stringer_base_t, stringer_neck_width=stringer_neck_width,
                            E=E, sf=sf,
                            sigma_e=E * np.square(np.pi) / (12 * (1 - np.square(mu))) * np.square(t / b))
    return read_excel(in_plane_stresses, axial_stringer_stresses, params=parameters), io, parameters


def fill_csv_with_commas(input_file):
    with open(input_file, 'r') as infile:
        reader = csv.reader(infile)
        rows = list(reader)
    max_delimiters = max(len(row) - 1 for row in rows)
    for row in rows:
        num_commas_needed = max_delimiters - (len(row) - 1)
        row.extend([''] * num_commas_needed)
    with open(input_file, 'w', newline='') as outfile:
        writer = csv.writer(outfile)
        writer.writerows(rows)


def read_excel(in_plane_stresses: str, axial_stringer_stresses: str, params: Parameters) -> [LoadCase]:
    fill_csv_with_commas(in_plane_stresses)
    fill_csv_with_commas(axial_stringer_stresses)
    logger.info('Reading excel file')
    ips = pd.read_csv(in_plane_stresses)
    ass = pd.read_csv(axial_stringer_stresses)
    load_cases = [LoadCase([], [], [], []) for _ in range(3)]

    k = 10
    j = 10
    for i, load_case in enumerate(load_cases):
        while int(ips.iat[k, 2]) == i + 1:
            load_case.LoadsInPlane.append(
                LoadsInPlane(e_id=int(ips.iat[k, 0]), xx=float(ips.iat[k, 5]),
                             yy=float(ips.iat[k, 7]),
                             xy=float(ips.iat[k, 6]),
                             von_Mises=float(ips.iat[k, 8]),
                             reserve_factor=np.abs(params.sigma_ul / (params.sf * float(ips.iat[k, 8])))))
            k += 1
            if k > len(ips) - 1:
                break
        while int(ass.iat[j, 2]) == i + 1:
            load_case.LoadsStringers.append(
                LoadsStringers(e_id=int(ass.iat[j, 0]), stress=float(ass.iat[j, 4]),
                               reserve_factor=np.abs(
                                   params.sigma_ul / (params.sf * float(ass.iat[j, 4])))))
            j += 1
            if j > len(ass) - 1:
                break
    logger.info('Finished reading excel file')
    return load_cases


def parse_excel(io: IO, load_cases: [LoadCase]):
    wb = openpyxl.load_workbook(io.output_file)
    ws = wb.get_sheet_by_name('in')
    parse_reserve_factors(ws, load_cases)
    parse_plane_analysis(ws, load_cases)
    parse_stringer_analysis(ws, load_cases)
    wb.save(io.output_file)
    logger.info('Finished parsing excel file')


def parse_reserve_factors(ws, load_cases):
    column = 2
    for i, load_case in enumerate(load_cases):
        row_loads_in_plane = 17
        for k, load in enumerate(load_case.LoadsInPlane):
            cellref = ws.cell(row=row_loads_in_plane + k, column=column)
            cellref.value = load.reserve_factor
        row_loads_stringers = 47
        for k, load in enumerate(load_case.LoadsStringers):
            cellref = ws.cell(row=row_loads_stringers + k, column=column)
            cellref.value = load.reserve_factor
        column += 3


def parse_plane_analysis(ws, load_cases):
    column = 2
    for i, load_case in enumerate(load_cases):
        row = 64
        for k, panel in enumerate(load_case.Panels):
            cellref = ws.cell(row=row + k, column=column)
            cellref.value = panel.avg_xx
            cellref = ws.cell(row=row + k, column=column + 1)
            cellref.value = panel.avg_yy
            cellref = ws.cell(row=row + k, column=column + 2)
            cellref.value = panel.avg_xy
            cellref = ws.cell(row=row + k, column=column + 3)
            cellref.value = panel.k_tau
            cellref = ws.cell(row=row + k, column=column + 4)
            cellref.value = panel.k_biax
            cellref = ws.cell(row=row + k, column=column + 5)
            cellref.value = panel.reserve_factor
        column += 8


def parse_stringer_analysis(ws, load_cases):
    column = 2
    for i, load_case in enumerate(load_cases):
        row = 73
        row_geo = 80
        for k, stringer in enumerate(load_case.Stringers):
            cellref = ws.cell(row=row + k, column=column)
            cellref.value = stringer.sigma_axial
            cellref = ws.cell(row=row + k, column=column + 1)
            cellref.value = stringer.sigma_crip
            cellref = ws.cell(row=row + k, column=column + 2)
            cellref.value = stringer.reserve_factor
            parse_geometric_properties(ws, stringer, row_geo + k)
        column += 5


def parse_geometric_properties(ws, stringer: Stringer, row):
    column = 2
    for k, geo in enumerate(stringer.geometrics):
        cellref = ws.cell(row=row, column=column)
        cellref.value = geo.I
        cellref = ws.cell(row=row, column=column + 1)
        cellref.value = geo.r_gyr
        cellref = ws.cell(row=row, column=column + 2)
        cellref.value = geo.lamda
        cellref = ws.cell(row=row, column=column + 3)
        cellref.value = geo.lamda_crit
